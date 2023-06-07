import os
import discord
from datetime import datetime, timedelta
import json
import pandas as pd
import shutil
from openpyxl import load_workbook
import random
import asyncio
import time

#Mine
from Recognition import IdentifyGame, getText
from api_LoL import getStats
from api_giphy import getGif

counter = {}
champMsg_id = 0 

# Ielādē nepieciešamo json failu
async def load_json_file(file_name):
    with open(file_name, 'r') as f:
        return json.load(f)

# Reģistrē uzvaru konkrētam spēlētājam
async def RegisterWin(wins, message, recap, sendConfirm, isStreak):
    print('RegisterWin initiated')

    # Ja ir , tad izdzēš ziņu ar bildi, kur ir lol čempions
    global champMsg_id
    try:
        message_obj = await message.channel.fetch_message(champMsg_id)
    except discord.NotFound:
        # message with given ID does not exist
        print(f"Message with ID {champMsg_id} does not exist")
    else:
        # message with given ID exists
        await message_obj.delete()
        print("LoL champion image deleted")

    mp_modes ={'HARDPOINT', 'SEARCH AND DESTROY', 'CONTROL', 'DOMINATION'}

    synonyms_impressive =   {}  
    synonyms_confirmed =    {}
    
    with open('synonyms_confirmed.json', 'r') as f:
       synonyms_confirmed = json.load(f)
    with open('synonyms_impressive.json', 'r') as f:
       synonyms_impressive = json.load(f)


    # Palielana uzvaru skaitu
    def increment_win(user_id, game_name):
        if user_id not in game_wins:
            game_wins[user_id] = {}
        if game_name not in game_wins[user_id]:
            game_wins[user_id][game_name] = 0
        game_wins[user_id][game_name] += 1

   
    def counterr(player):
        print('Counter initiated')
        global counter

        if player not in counter:
            counter[player] = {}
            counter[player] = 3
        counter[player] +=1
        return counter

    # Iegūst lietotāju un uzvaru datubāzi
    user_id = str(message.author)

    game_wins = wins

    role_id = 1030500402023628850  
    role = message.guild.get_role(role_id)
    await message.author.add_roles(role)

    # Save image
    for attachment in message.attachments:
        timee = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        filename = f"{message.author.discriminator}_{timee}_{attachment.filename}"
        await attachment.save(f'./attachments/{filename}')
        print(f'Saglabāju bildi {filename} no {message.author}')

    # Increment the game wins for the user and game
    game_name = IdentifyGame(f'C:/Users/ZK00138/source/repos/ResnsCounter/ResnsCounter/attachments/{filename}')
    if game_name != 'Unrecognised':

        increment_win(user_id, game_name)
    
        msg_time = message.created_at
        specifDay = msg_time.strftime('%A')
        specifTime = message.created_at.hour
        specifDate = message.created_at.day

        await regSpecifDay(specifDay, user_id)
        await regDayTime(specifTime,user_id)
        await regEachDay(specifDate,user_id)


    # Respond with a confirmation message
        await message.add_reaction('✅')
        response = ""
#################################################### UPDATED
        if message.mentions:
            mentioned_users = set()  # set to keep track of unique users mentioned
            for user in message.mentions:
                mention_id = f'{user.name}#{user.discriminator}'
                if mention_id not in mentioned_users:  # check if user has already been mentioned
                    mentioned_users.add(mention_id)
                    increment_win(mention_id, game_name)
                    await regDayTime(specifTime, mention_id)
                    await regSpecifDay(specifDay, mention_id)
                    await regEachDay(specifDate,mention_id)
                    #await RegTotalMonthWins(1, message.created_at.month)
#################################################### UPDATED

                if not recap:
                    response = response + f"{user.mention} " #Tev tagad ir {game_wins[mention_id][game_name]} uzvaras."

                # Pievieno role lietotājam
                await user.add_roles(role)
        if not recap:


            if sendConfirm:
                syno_conf = random.choice(list(synonyms_confirmed.values())) 
                if not isStreak: response = response + f"{message.author.mention} {game_name} win {syno_conf}!" #At this moment you have {game_wins[user_id][game_name]} victories."
                if isStreak:
                   synonyms_WinnStreak = await load_json_file('synonyms_WinnStreak.json')
                   syno_winStreak = random.choice(list(synonyms_WinnStreak.values()))
                   response = response + f"{message.author.mention} {game_name} wins confirmed! {syno_winStreak}"
########################################### UPDATED ################################
                extra =""
                title =""
                phrase = "default"
                if game_name == "League of Legends":
                    phrases_LoL = await load_json_file('phrases_LoL.json')
                    #gifs_LoL =    await load_json_file('gifs_LoL.json')

                    if not isStreak:
                        champ_name, num_kills = getStats(user_id) 
                        if champ_name is not None and num_kills is not None:
                            num_kills = str(num_kills)
                            disp_name = message.author.display_name
                            syno_impresive = random.choice(list(synonyms_impressive.values()))
                            stats = f"\n\n{syno_impresive} **{num_kills}** kills with **{champ_name}**, {disp_name}!"
                            response = response + stats
                            file_name = os.path.join("images_LoL", f"{champ_name}.png")
                            with open(file_name, "rb") as f:
                                file = discord.File(f)

                    phrase = random.choice(list(phrases_LoL.values()))
                    gif =    getGif("League of Legends")

                    await message.channel.send(response)
                    counterr(user_id)
                    if counter[user_id] %4 == 0:
                        await message.channel.send(phrase)
                     # Aizsūta un pēc laika izdzēš gif. Champion bilde tiks izdzēsta, kad tiks reģistrēta nākošā iesūtītā uzvara
                    if not isStreak:                
                       try: #@#
                           img_msg = await message.channel.send(file=file) #@#
                           champMsg_id = img_msg.id #@#
                       except: #@#
                           pass
                    gif_msg = await message.channel.send(gif)
                    await asyncio.sleep(20)
                    await gif_msg.delete()
                    

                elif game_name == "COD MW2 Multiplayer":
                    #Ielādē gif
                    #gifs_mp = await load_json_file('gifs_multiplayer.json')

                    for mode in mp_modes:
                        result = getText(f'C:/Users/ZK00138/source/repos/ResnsCounter/ResnsCounter/attachments/{filename}', mode, mode)
                        if result is not None:
                            exist, title = result
                    extra = title
                    if   extra == "HARDPOINT":
                        phrases_mp_hardpoint = await load_json_file('phrases_mp_hardpoint.json')
                        phrase = random.choice(list(phrases_mp_hardpoint.values()))

                    elif extra == "CONTROL":
                        phrases_mp_control = await load_json_file('phrases_mp_control.json')
                        phrase = random.choice(list(phrases_mp_control.values()))

                    elif extra == "SEARCH AND DESTROY":
                        phrases_mp_SnD = await load_json_file('phrases_mp_SnD.json')
                        phrase = random.choice(list(phrases_mp_SnD.values()))

                    elif extra == "DOMINATION":
                        phrases_mp_domination = await load_json_file('phrases_mp_domination.json')
                        phrase = random.choice(list(phrases_mp_domination.values()))

                    else:
                        phrases_mp = await load_json_file('phrases_multiplayer.json')
                        phrase = random.choice(list(phrases_mp.values()))

                    gif =    getGif("call of duty multiplayer") 
                    counterr(user_id)
                    await message.channel.send(response)  
                    # Papildus paziņojumu izsūta ik pēc 3 uzvarām
                    if counter[user_id] %4 == 0:
                        await message.channel.send(phrase)
                    gif_msg = await message.channel.send(gif)
                    await asyncio.sleep(20)
                    await gif_msg.delete()


                elif game_name == "Warzone 2.0" or game_name == "Warzone 1.0":
                    phrases_warzone = await load_json_file('phrases_warzone.json')
                    #gifs_warzone =    await load_json_file('gifs_warzone.json')
                    random.seed(time.time())
                    phrase = random.choice(list(phrases_warzone.values()))
                    gif =    getGif("Warzone") 
                    await message.channel.send(response)
                    counterr(user_id)
                    if counter[user_id] %4 == 0:
                        await message.channel.send(phrase)
                    gif_msg = await message.channel.send(gif)
                    await asyncio.sleep(20)
                    await gif_msg.delete()

                elif game_name == "CSGO": 
                    phrases_CSGO = await load_json_file('phrases_CSGO.json')
                    #gifs_CSGO = await load_json_file('gifs_CSGO.json')
                    random.seed(time.time())
                    phrase = random.choice(list(phrases_CSGO.values()))
                    gif =    getGif("CSGO")
                    await message.channel.send(response)
                    await message.channel.send(phrase)
                    gif_msg = await message.channel.send(gif)
                    await asyncio.sleep(20)
                    await gif_msg.delete()

                elif game_name == "Destiny 2": 
                    phrases_Destiny2 = await load_json_file('phrases_Destiny2.json')
                    #gifs_Destiny2 =    await load_json_file('gifs_Destiny2.json')
                    random.seed(time.time())
                    phrase = random.choice(list(phrases_Destiny2.values())) 
                    gif =    getGif("Destiny 2")
                    await message.channel.send(response)
                    await message.channel.send(phrase)
                    gif_msg = await message.channel.send(gif)
                    await asyncio.sleep(20)
                    await gif_msg.delete()

                elif game_name == "Apex Legends": 
                    phrases_Apex_Legends = await load_json_file('phrases_Apex_Legends.json')
                    #gifs_Apex_Legends =    await load_json_file('gifs_Apex_Legends.json')
                    random.seed(time.time())
                    phrase = random.choice(list(phrases_Apex_Legends.values())) 
                    gif =    getGif("Apex Legends")
                    await message.channel.send(response)
                    await message.channel.send(phrase)
                    gif_msg = await message.channel.send(gif)
                    await asyncio.sleep(20)
                    await gif_msg.delete() 

                elif game_name == "Fortnite":  
                    phrases_Fortnite = await load_json_file('phrases_Fortnite.json')
                    #gifs_Fortnite =    await load_json_file('gifs_Fortnite.json')
                    random.seed(time.time())
                    phrase = random.choice(list(phrases_Fortnite.values())) 
                    gif =    getGif("Fortnite")
                    await message.channel.send(response)
                    await message.channel.send(phrase)
                    gif_msg = await message.channel.send(gif)
                    await asyncio.sleep(20)
                    await gif_msg.delete()

                elif game_name == "War Thunder":  
                    gif = getGif("War Thunder")
                    await message.channel.send(response)
                    gif_msg = await message.channel.send(gif)
                    await asyncio.sleep(20)
                    await gif_msg.delete()

        # Saglaba json failā
        with open('resnums.json', 'w') as f:
            json.dump(game_wins, f)

        # Saglabā excel failā no json
        df = pd.read_json('resnums.json')
        df = df.transpose()
        df.to_excel('resnums.xlsx', index=True)

        # Formatē tabulu
        workbook = load_workbook('resnums.xlsx')
        # Get the first sheet in the workbook
        sheet = workbook['Sheet1']
        # Set the width of the first column to 25
        sheet.column_dimensions['A'].width = 35
        for column in ['B', 'C', 'D','E', 'F','G']:
            sheet.column_dimensions[column].width = 20
        # Save the changes to the Excel file
        workbook.save('resnums.xlsx')


    else:
        #Neatpazits attēls
        file_path = f'C:/Users/ZK00138/source/repos/ResnsCounter/ResnsCounter/attachments/{filename}'
        destination_folder_path = 'C:/Users/ZK00138/source/repos/ResnsCounter/ResnsCounter/Neatpazitie'
        shutil.copy(file_path, destination_folder_path)
        print('Pārvietoju bildi, kura netika atpazīta')
        if not recap:
            response = f"{message.author.mention} Ooh noo, mommy I need help, I did not recognize  the game. Anyhow, congrats. It will be added manually"
            await message.channel.send(response)

    return wins

# Reģistrē kurā nedēļas dienā tika gūta uzvara
async def regSpecifDay(day, player):

    SpecifDay = {}

    filename = "SpecifDay.json"
    if os.path.isfile(filename):
       with open('SpecifDay.json', 'r') as f:
           SpecifDay = json.load(f)  

    if day not in SpecifDay:
        SpecifDay[day] = {}
    if player not in SpecifDay[day]:
        SpecifDay[day][player] = 0
    SpecifDay[day][player] += 1

    with open('SpecifDay.json', 'w') as f:
        json.dump(SpecifDay, f)

    # Saglabā excel failā no json
    df = pd.read_json('SpecifDay.json')
    df.to_excel('SpecifDay.xlsx', index=True)

    # Formatē tabulu
    workbook = load_workbook('SpecifDay.xlsx')
    # Get the first sheet in the workbook
    sheet = workbook['Sheet1']
    # Set the width of the first column to 25
    sheet.column_dimensions['A'].width = 35
    for column in ['A','B', 'C', 'D', 'E', 'F', 'G', 'H']:
        sheet.column_dimensions[column].width = 20
    # Save the changes to the Excel file
    workbook.save('SpecifDay.xlsx')

# Reģistrē kurā dienas daļā gūta uzvara
async def regDayTime(hour, player):
    Period = ""
    DayTime = {}

    hour = hour + 3
    if hour > 24: hour = hour - 24

    filename = "DayTime.json"
    if os.path.isfile(filename):
       with open('DayTime.json', 'r') as f:
           DayTime = json.load(f)  



    if 6 <= hour and hour <= 12:
        Period = "Mornning(6-12)"
    elif 12 <= hour and hour <= 18:
        Period = "Afternoon(12-18)"
    elif 18 <= hour and hour <= 24:
        Period = "Evening(18-24)"
    else: Period = "Night(24-6)"

    if Period not in DayTime:
        DayTime[Period] = {}
    if player not in DayTime[Period]:
        DayTime[Period][player] = 0
    DayTime[Period][player] += 1

    with open('DayTime.json', 'w') as f:
        json.dump(DayTime, f)

    # Saglabā excel failā no json
    df = pd.read_json('DayTime.json')
    df.to_excel('DayTime.xlsx', index=True)

    # Formatē tabulu
    workbook = load_workbook('DayTime.xlsx')
    # Get the first sheet in the workbook
    sheet = workbook['Sheet1']
    # Set the width of the first column to 25
    sheet.column_dimensions['A'].width = 35
    for column in ['A','B', 'C', 'D', 'E']:
        sheet.column_dimensions[column].width = 20
    # Save the changes to the Excel file
    workbook.save('DayTime.xlsx')

# Reģistrē kurā stundā gūta uzvara
async def Register_time(hour):

    
    #Ielādē datus no faila
    filename = "laiks.json"
    if os.path.isfile(filename):
       with open('laiks.json', 'r') as f:
           time_statistics = json.load(f)
    else: time_statistics = {}
    if int(hour) > 24: hour = int(hour) - 24
    hour = str(hour)
    hour = f'{hour}: 00'
    #Parbauda vai konkrētā stunda jau nav reģistrēta
    if hour not in time_statistics:
        time_statistics[hour] = {}
        time_statistics[hour] = 0

    time_statistics[hour]+= 1
    

    # Saglaba json failā
    with open('laiks.json', 'w') as f:
        json.dump(time_statistics, f)


    # Saglabā excel failā no json
    df = pd.read_json('laiks.json', typ='series', orient='index')
    df = df.transpose()
    df.to_excel('laiks.xlsx', index=True)

    return time_statistics

 # Reģistrē kopējo uzvaru skaitu mēnesī
async def RegTotalMonthWins(amount, month):

   winsInMonth =        {}
   month = f'{month}_2023'
   with open('winsInMonth.json', 'r') as f:
       winsInMonth = json.load(f)

   if month not in winsInMonth:
        winsInMonth[month] = {}
        winsInMonth[month] = 0

   winsInMonth[month] = winsInMonth[month] + amount
   with open('winsInMonth.json', 'w') as f:
        json.dump(winsInMonth, f)

################# NEW ###################
async def regEachDay(day, player):
    
    EachDay = {}

    filename = "EachDay.json"
    if os.path.isfile(filename):
       with open('EachDay.json', 'r') as f:
           EachDay = json.load(f)  
    dayy = str(day)

    if dayy not in EachDay:
        EachDay[dayy] = {}
    if player not in EachDay[dayy]:
        EachDay[dayy][player] = 0
    EachDay[dayy][player] += 1

    with open('EachDay.json', 'w') as f:
        json.dump(EachDay, f)

    # Saglabā excel failā no json
    df = pd.read_json('EachDay.json')
    df.to_excel('EachDay.xlsx', index=True)

    # Formatē tabulu
    workbook = load_workbook('EachDay.xlsx')
    # Get the first sheet in the workbook
    sheet = workbook['Sheet1']
    # Set the width of the first column to 25
    sheet.column_dimensions['A'].width = 25
    # Save the changes to the Excel file
    workbook.save('EachDay.xlsx')
################# NEW ###################
